import datetime
import json
import string
import random
import openpyxl
import pandas as pd
import requests
from openpyxl.styles import PatternFill

from config.config import *
from libraries.logger import logg


def funName(fucntioname):
    # lstr = fucntioname.split('::')
    # fun = lstr[2]
    m = fucntioname.find('Dependent')
    if m > 0:
        nam = fucntioname[m + 11:]
    else:
        nam = fucntioname
    return nam


def get_random_string(length):
    # choose from all lowercase letter
    letters = string.ascii_lowercase
    result_str = ''.join(random.choice(letters) for i in range(length))
    print("Random string of length", length, "is:", result_str)
    return result_str


def fn_Postrequest(testcasename, endpt):
    regobj, data = getdata(testcasename, endpt)
    url = baseURL + data[endpt]
    logg.info("End point URL is :%s" % url)
    logg.info("Header is :%s" % headers)
    logg.info("Post body is :%s" % regobj)
    reponse = requests.post(url, json=regobj, headers=headers)
    # reponse=reponse.json()
    return reponse, regobj, data

def fn_Getrequest(testcasename, endpt):
    regobj, data = getdata(testcasename, endpt)
    url = baseURL + data[endpt]
    if endpt == 'user_url':
        getheaders={}
        getheaders['Authorization'] = 'Bearer ' + data['Token']
        url = url + str(data['Id'])
    logg.info("This is a Get Request")
    logg.info("End point URL is :%s" % url)
    logg.info("Header is :%s" % getheaders)
    reponse = requests.get(url, headers=getheaders)
    # reponse=reponse.json()
    return reponse, regobj, data

def verifyUserCreattion(testcasename):
    logg.info("Testcase name %s:" % testcasename)
    tccasename = funName(testcasename)
    res, upobj, exclobj = fn_Postrequest(tccasename, 'regi_url')
    if res.status_code == 200:
        logg.info("Success : Successfully Verified Response from API is %s" % res)
        res = res.json()
        update_data(tccasename, res, upobj)
        strmsg="Success : Passed : Successfully Registered user"
        logg.info(strmsg)
        logg.info("Success : Response from APi : %s" % res)
        update_Result(testcasename,'Passed',strmsg)
    else:
        logg.error("Failed : Not able to Register the user and Response from APi%s" % res.json())


def loginwiththeregistereduser(testcasename):
    logg.info("Test case Execution Started for %s" % testcasename)
    tccasename = funName(testcasename)
    res, upobj, exclobj = fn_Postrequest(tccasename, 'login_url')
    if res.status_code == 200:
        logg.info("Success : Successfully Verified Response from API is %s" % res)
        res = res.json()
        if VerifydatainReponse(testcasename, res, exclobj):
            strmsg="Success : Successfully logged in and  Verified ID, Name and Email id from API and provided during " \
                   "Registration "
            update_Result(testcasename, 'Passed', strmsg)
            logg.info(strmsg)
        else:
            logg.info("Failed : Mismatch Data, Name and Email id from API is not same and provided durin")
    else:
        logg.error("Failed : Not able to login with user and Response from APi%s" % res.json())


def getUSerbyID(testcasename):
    logg.info("Test case Execution Started for %s" % testcasename)
    tccasename = funName(testcasename)
    res, upobj, exclobj = fn_Getrequest(tccasename, 'user_url')
    if res.status_code == 200:
        logg.info("Success : Successfully Verified Response from API is %s" % res)
        if VerifydatainReponse(testcasename, res, exclobj):
            strmsg="Success : Successfully Verified ID, Name and Email id from API: and provided during Registration"
            update_Result(testcasename, 'Passed', strmsg)
            logg.info(strmsg)
        else:
            strmsg="Failed : Mismatch Data, Name and Email id from API is not same and provided during Registration"
            update_Result(testcasename, 'Failed', strmsg)
            logg.error(strmsg)

    else:
        logg.error("Failed: Response code from API is %s" % res)
        logg.error("Failed: Not able to Hit API  and Response from APi :%s" % res.json())
        strmsg = "Failed: Not able to Hit API Response code is %s and  Response from API is  %s" % (res,res.json())
        update_Result(testcasename, 'Failed', strmsg)
        #logg.error(strmsg)

def getdata(tc_name, endpt):
    try:
        obj = {}
        d = {}
        dataregi = {}
        datalogin = {}
        datadf = pd.read_excel('../test_data/TestData.xlsx')
        # print(datadf)
        s = datadf[datadf['TestCaseName'] == tc_name].index[0]
        headername = list(datadf.columns.values)
        print(headername)
        for i in range(datadf.shape[1]):
            d[headername[i]] = datadf.iat[s, i]
        # print(d)
    except Exception as er:
        print('something is wrong ')
    if endpt == 'regi_url':
        d['name'] = dataregi['name'] = d['name'] + get_random_string(4)
        d['email'] = dataregi['email'] = d['email'] + get_random_string(4)
        dataregi['password'] = d['password']
        obj = dataregi
    elif endpt == 'login_url':
        datalogin['email'] = d['email']
        datalogin['password'] = d['password']
        obj = datalogin
    elif endpt == 'user_url':
        obj
    return obj, d


def update_data(tc_name, res, upobj):
    wb_obj = openpyxl.load_workbook('../test_data/TestData.xlsx')
    sheet_obj = wb_obj.active
    max_column = sheet_obj.max_column
    max_row = sheet_obj.max_row

    for x in range(1, max_row + 1):
        if sheet_obj.cell(row=x, column=1).value == tc_name:
            sheet_obj.cell(row=x, column=5).value = res['data']['Token']  # Token
            sheet_obj.cell(row=x, column=6).value = res['data']['Id']  # id
            sheet_obj.cell(row=x, column=2).value = upobj['name']
            sheet_obj.cell(row=x, column=3).value = upobj['email']
            # sheet_obj.cell(row=x, column=7).value='Test3'

    wb_obj.save('../test_data/TestData.xlsx')


def update_Result(tc_name, exeStatus, strmsg):
    wb_obj = openpyxl.load_workbook('../result/Result.xlsx')
    sheet_obj = wb_obj.active
    max_column = sheet_obj.max_column
    max_row = sheet_obj.max_row
    fill_green = PatternFill(patternType='solid',
                             fgColor='35FC03')
    fill_red = PatternFill(patternType='solid',
                             fgColor='FC2C03')
    x = max_row+1
    sheet_obj.cell(row=x, column=1).value = tc_name  # TestCaseName
    sheet_obj.cell(row=x, column=2).value = datetime.datetime.now()  # #Timestamp
    if exeStatus=='Passed':
        sheet_obj.cell(row=x, column=3).fill=fill_green
        sheet_obj.cell(row=x, column=3).value = exeStatus  # Execution_Status
    else:
        sheet_obj.cell(row=x, column=3).fill=fill_red
        sheet_obj.cell(row=x, column=3).value = exeStatus
    sheet_obj.cell(row=x, column=4).value = strmsg
    wb_obj.save('../result/Result.xlsx')


def VerifydatainReponse(testcasename, res, exclobj):
    if res['data']['Id'] == exclobj['Id'] and res['data']['Name'] == exclobj['name'] and res['data']['Email'] == \
            exclobj['email']:
        logg.info("Success : Successfully compared Id from API:%s   and provided during Registration: %s" % (
            res['data']['Id'], exclobj['Id']))
        logg.info("Success : Successfully compared Name from API: %s   and provided during Registration: %s" % (
            res['data']['Name'], exclobj['name']))
        logg.info("Success : Successfully compared Email id from API: %s  and provided during Registration: %s" % (
            res['data']['Email'], exclobj['email']))
        return True
    else:
        logg.error(
            "Failed :Mismatch Value are  Id ,name and ID from API %s %s %s from API  and provided during Registration  "
            "are "
            "not same%s %s %s " % (
                res['data']['Id'], res['data']['Name'], res['data']['Email'], exclobj['Id'], exclobj['name'],
                exclobj['email']))
        return False
